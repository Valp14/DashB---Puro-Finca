"""
pages/reportes.py
-----------------
Reportes filtrable por proceso con exportacion a CSV, Excel simple o
Excel consolidado ejecutivo de 8 hojas.
[Equivalente a pages/10_Reportes.py de Streamlit]

[MIGRACION Streamlit -> Dash]
Streamlit usaba st.download_button con `data=bytes`. Aqui usamos
dcc.Download (ya definidos en layout.py como download-csv, download-xlsx,
download-xlsx-consolidado) + dcc.send_bytes / dcc.send_data_frame.

Los 3 botones son dbc.Button controlados con n_clicks + callback_context
para saber cual se presiono.
"""

from __future__ import annotations

import io
from datetime import date

import pandas as pd
import dash_bootstrap_components as dbc
from dash import Input, Output, State, callback, ctx, dcc, html

from config.settings import (PROCESOS, PROCESOS_LABEL, ESTANDARES,
                              DOTACION_ESTANDAR,
                              UMBRAL_CUMPLIMIENTO_OK, UMBRAL_CUMPLIMIENTO_ALERTA,
                              UMBRAL_PRIMERA_OK, UMBRAL_PRIMERA_ALERTA,
                              UMBRAL_COMERCIAL_OK, UMBRAL_COMERCIAL_ALERTA,
                              UMBRAL_PERDIDA_OK, UMBRAL_PERDIDA_ALERTA)
from styles.theme import page_header, empty_state
from pages._common import filter_panel
from components.ui import detail_table, fmt_num, kpi_card, kpi_row
from components.sidebar import get_data_filtrada, store_to_dataframes
from utils.metrics import (kpis_corte, kpis_siembra, kpis_cosecha,
                            kpis_lavado, kpis_empaque, kpis_cargue,
                            analisis_dotacion, detectar_inconsistencias)


# ---------------------------------------------------------------------------
# Layout
# ---------------------------------------------------------------------------
def layout():
    proceso_labels = [PROCESOS_LABEL[p] for p in PROCESOS]

    return html.Div([
        page_header("Reportes",
                    subtitle="Consulta detallada, filtros y exportaciones"),

        # Sin archivo -> mensaje
        html.Div(id="reportes-gate"),
        filter_panel(),

        # Controles
        html.Div([
            html.Div([
                html.Div("Proceso a consultar", className="control-label"),
                dcc.Dropdown(
                    id="reporte-proceso",
                    options=[{"label": lbl, "value": lbl} for lbl in proceso_labels],
                    value=proceso_labels[2],  # Cosecha por defecto
                    clearable=False,
                ),
            ]),
            html.Div([
                html.Div("Buscar en observaciones (opcional)",
                         className="control-label"),
                dcc.Input(id="reporte-busqueda", type="text",
                          placeholder="Ej: maquinaria, lluvia, retraso",
                          style={"width": "100%", "padding": "8px",
                                 "border": "1px solid #E3E6EB",
                                 "borderRadius": "4px"}),
            ]),
        ], className="two-col",
           id="reportes-controles"),

        html.Div(id="reportes-kpis"),
        html.Hr(),

        html.H2("Tabla de detalle"),
        html.Div(id="reportes-tabla"),

        html.H2("Exportar"),
        html.Div(id="reportes-exportar"),

        html.H2("Vista consolidada por finca y lote"),
        html.Div(id="reportes-consolidado"),
    ], className="ops-page ops-process-page process-reportes")


# ---------------------------------------------------------------------------
# Callback principal: renderiza todo el contenido visible
# ---------------------------------------------------------------------------
@callback(
    Output("reportes-gate",        "children"),
    Output("reportes-kpis",        "children"),
    Output("reportes-tabla",       "children"),
    Output("reportes-exportar",    "children"),
    Output("reportes-consolidado", "children"),
    Input("store-archivo",      "data"),
    Input("store-data",         "data"),
    Input("store-filtros",      "data"),
    Input("reporte-proceso",    "value"),
    Input("reporte-busqueda",   "value"),
)
def render(archivo, store_data, filtros, proceso_lbl, busqueda):
    if not archivo:
        gate = empty_state("No se encontraron datos en Supabase para generar los reportes.")
        return gate, None, None, None, None

    proceso_label_a_canonical = {v: k for k, v in PROCESOS_LABEL.items()}
    proceso = proceso_label_a_canonical.get(proceso_lbl, PROCESOS[0])

    df = get_data_filtrada(store_data, filtros, proceso)

    # Filtro local por observaciones
    if busqueda and not df.empty and "Observaciones" in df.columns:
        mask = df["Observaciones"].fillna("").astype(str).str.contains(
            busqueda, case=False, na=False, regex=False)
        df_view = df[mask]
    else:
        df_view = df

    # KPIs rapidos
    fincas_distintas = (df_view["Finca"].nunique()
                        if "Finca" in df_view.columns else 0)
    if "Fecha" in df_view.columns and not df_view.empty:
        s = pd.to_datetime(df_view["Fecha"], errors="coerce").dropna()
        rng = (f"{s.min().date()} → {s.max().date()}" if not s.empty else "-")
    else:
        rng = "-"
    kpis = kpi_row([
        kpi_card("Registros mostrados", fmt_num(len(df_view)),
                 sub="Filtros aplicados", nivel="neutral"),
        kpi_card("Fincas", fmt_num(fincas_distintas),
                 sub="Con registros en la vista", nivel="neutral"),
        kpi_card("Rango de fechas", rng, sub="Min → Max", nivel="neutral"),
    ])

    # Tabla
    if df_view.empty:
        tabla = empty_state("Sin registros para los filtros y búsqueda actuales.")
    else:
        show = df_view.copy()
        for c in show.columns:
            if pd.api.types.is_datetime64_any_dtype(show[c]):
                show[c] = show[c].dt.strftime("%Y-%m-%d %H:%M")
        tabla = detail_table(show, height=420, id="reporte-tabla-detalle")

    # Botones de exportar
    if df_view.empty:
        exp = html.Div("No hay datos para exportar.", className="muted")
    else:
        exp = html.Div([
            html.Div(
                dbc.Button("Descargar CSV (proceso actual)",
                           id="btn-download-csv", n_clicks=0,
                           color="primary", outline=True,
                           style={"width": "100%"})
            ),
            html.Div(
                dbc.Button("Descargar Excel (proceso actual)",
                           id="btn-download-xlsx", n_clicks=0,
                           color="primary", outline=True,
                           style={"width": "100%"})
            ),
            html.Div(
                dbc.Button("Descargar Excel consolidado (8 hojas)",
                           id="btn-download-consol", n_clicks=0,
                           color="primary",
                           style={"width": "100%"},
                           title="Excel ejecutivo con resumen, productividad, "
                                 "calidad, dotación, plan vs real, detalle "
                                 "por proceso, calidad de datos y parámetros.")
            ),
        ], style={"display": "grid",
                  "gridTemplateColumns": "repeat(3, 1fr)",
                  "gap": "12px"})

    # Vista consolidada por finca/lote
    agregables = {
        "Corte Esquejes":       ["Esquejes", "Horas"],
        "Siembra":              ["Plantas", "Horas"],
        "Cosecha":              ["Produccion Kg", "Descarte Kg", "Horas"],
        "Lavado Clasificacion": ["Kg Recibidos", "Kg Lavados", "Kg 1ra",
                                 "Kg 2da", "Kg 3ra", "Kg Semilla",
                                 "Kg Descarte Lavado", "Horas"],
        "Empaque":              ["Kg Recibidos", "Kg Empacados",
                                 "Kg Cajas", "Kg Sacos", "Kg Bolsas", "Horas"],
        "Cargue Vehiculo":      ["Toneladas", "Cajas", "Sacos", "Bolsas", "Horas"],
    }
    cols_agreg = [c for c in agregables.get(proceso, []) if c in df_view.columns]
    group_cols = [c for c in ("Finca", "Lote") if c in df_view.columns]

    if cols_agreg and group_cols and not df_view.empty:
        consolidado = (df_view.groupby(group_cols, dropna=False)[cols_agreg]
                              .sum(numeric_only=True)
                              .reset_index())
        cons_out = detail_table(consolidado, id="tbl-consolidado-fl")
    else:
        cons_out = empty_state("No hay columnas agregables para consolidar en este proceso.")

    return None, kpis, tabla, exp, cons_out


# ---------------------------------------------------------------------------
# Callback: descargar CSV (reemplaza st.download_button CSV)
# ---------------------------------------------------------------------------
@callback(
    Output("download-csv", "data"),
    Input("btn-download-csv", "n_clicks"),
    State("reporte-proceso",  "value"),
    State("reporte-busqueda", "value"),
    State("store-data",       "data"),
    State("store-filtros",    "data"),
    prevent_initial_call=True,
)
def descargar_csv(n, proceso_lbl, busqueda, store_data, filtros):
    if not n:
        return None
    proceso_label_a_canonical = {v: k for k, v in PROCESOS_LABEL.items()}
    proceso = proceso_label_a_canonical.get(proceso_lbl, PROCESOS[0])
    df = get_data_filtrada(store_data, filtros, proceso)
    if busqueda and not df.empty and "Observaciones" in df.columns:
        mask = df["Observaciones"].fillna("").astype(str).str.contains(
            busqueda, case=False, na=False, regex=False)
        df = df[mask]
    if df.empty:
        return None
    return dcc.send_data_frame(
        df.to_csv, f"reporte_{proceso.lower().replace(' ', '_')}_{date.today()}.csv",
        index=False, encoding="utf-8-sig",
    )


# ---------------------------------------------------------------------------
# Callback: descargar Excel simple
# ---------------------------------------------------------------------------
@callback(
    Output("download-xlsx", "data"),
    Input("btn-download-xlsx", "n_clicks"),
    State("reporte-proceso",   "value"),
    State("reporte-busqueda",  "value"),
    State("store-data",        "data"),
    State("store-filtros",     "data"),
    prevent_initial_call=True,
)
def descargar_xlsx(n, proceso_lbl, busqueda, store_data, filtros):
    if not n:
        return None
    proceso_label_a_canonical = {v: k for k, v in PROCESOS_LABEL.items()}
    proceso = proceso_label_a_canonical.get(proceso_lbl, PROCESOS[0])
    df = get_data_filtrada(store_data, filtros, proceso)
    if busqueda and not df.empty and "Observaciones" in df.columns:
        mask = df["Observaciones"].fillna("").astype(str).str.contains(
            busqueda, case=False, na=False, regex=False)
        df = df[mask]
    if df.empty:
        return None

    def _write(buffer):
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            exp = df.copy()
            for c in exp.columns:
                if pd.api.types.is_datetime64_any_dtype(exp[c]):
                    try:
                        exp[c] = exp[c].dt.tz_localize(None) if exp[c].dt.tz else exp[c]
                    except Exception:
                        pass
            exp.to_excel(writer, index=False, sheet_name=proceso[:31])

    return dcc.send_bytes(_write,
        f"reporte_{proceso.lower().replace(' ', '_')}_{date.today()}.xlsx")


# ---------------------------------------------------------------------------
# Callback: descargar Excel consolidado (8 hojas)
# ---------------------------------------------------------------------------
@callback(
    Output("download-xlsx-consolidado", "data"),
    Input("btn-download-consol", "n_clicks"),
    State("store-data",    "data"),
    State("store-filtros", "data"),
    prevent_initial_call=True,
)
def descargar_consolidado(n, store_data, filtros):
    if not n:
        return None

    def _write(buffer):
        from openpyxl.styles import Font, PatternFill, Alignment

        data = {p: get_data_filtrada(store_data, filtros, p) for p in PROCESOS}
        kcor = kpis_corte(data["Corte Esquejes"])
        ksie = kpis_siembra(data["Siembra"])
        kcos = kpis_cosecha(data["Cosecha"])
        klav = kpis_lavado(data["Lavado Clasificacion"])
        kemp = kpis_empaque(data["Empaque"])
        kcar = kpis_cargue(data["Cargue Vehiculo"])
        meta_ton_periodo = kcos["meta_ton_periodo"]

        def _round_or_none(value, digits=0):
            return round(value, digits) if value is not None else None

        def _pct_or_none(value, digits=1):
            return round(value * 100, digits) if value is not None else None

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # --- 1. Resumen -----------------------------------------------
            resumen = pd.DataFrame([
                {"Indicador": "Esquejes cortados (total)",     "Valor": kcor["total_esquejes"], "Unidad": "esquejes"},
                {"Indicador": "Plantas sembradas (total)",     "Valor": ksie["total_plantas"], "Unidad": "plantas"},
                {"Indicador": "Kilogramos cosechados (total)", "Valor": kcos["total_kg"], "Unidad": "kg"},
                {"Indicador": "Toneladas producidas",          "Valor": round(kcos["total_kg"]/1000, 2), "Unidad": "t"},
                {"Indicador": "Kg lavados (total)",            "Valor": klav["total_lavado"], "Unidad": "kg"},
                {"Indicador": "Kg empacados (total)",          "Valor": kemp["total_empacado"], "Unidad": "kg"},
                {"Indicador": "Toneladas despachadas",         "Valor": round(kcar["total_toneladas"], 2), "Unidad": "t"},
                {"Indicador": "Primera calidad",               "Valor": _pct_or_none(klav["pct_1"], 1), "Unidad": "%"},
                {"Indicador": "Segunda calidad",               "Valor": _pct_or_none(klav["pct_2"], 1), "Unidad": "%"},
                {"Indicador": "Tercera calidad",               "Valor": _pct_or_none(klav["pct_3"], 1), "Unidad": "%"},
                {"Indicador": "Caja comercializable (1ra+2da+3ra)", "Valor": _pct_or_none(klav["pct_comercial"], 1), "Unidad": "%"},
                {"Indicador": "Pérdida en lavado",             "Valor": _pct_or_none(klav["pct_perdida"], 2), "Unidad": "%"},
                {"Indicador": "Descarte en cosecha",           "Valor": _pct_or_none(kcos["pct_descarte"], 2), "Unidad": "%"},
                {"Indicador": "Toneladas promedio por despacho", "Valor": _round_or_none(kcar["ton_promedio"], 2), "Unidad": "t"},
                {"Indicador": "Cumplimiento cargue",           "Valor": _pct_or_none(kcar["cumpl_ton"], 1), "Unidad": "%"},
            ])
            resumen.to_excel(writer, sheet_name="1. Resumen", index=False)

            # --- 2. Productividad ----------------------------------------
            prod = pd.DataFrame([
                {"Proceso": "Corte de esquejes",    "Plan": kcor["estandar_persona"], "Real": _round_or_none(kcor["prod_persona_dia"]), "Unidad": "esquejes/persona/día", "Cumplimiento %": _pct_or_none(kcor["cumpl_prod"], 1)},
                {"Proceso": "Siembra",              "Plan": ksie["estandar_persona"], "Real": _round_or_none(ksie["prod_persona_dia"]), "Unidad": "plantas/persona/día",  "Cumplimiento %": _pct_or_none(ksie["cumpl_prod"], 1)},
                {"Proceso": "Cosecha",              "Plan": kcos["estandar_persona"], "Real": _round_or_none(kcos["prod_persona_dia"]), "Unidad": "kg/persona/día",       "Cumplimiento %": _pct_or_none(kcos["cumpl_prod"], 1)},
                {"Proceso": "Lavado y clasificación","Plan": klav["estandar_persona"],"Real": _round_or_none(klav["prod_persona_dia"]), "Unidad": "kg/persona/día",       "Cumplimiento %": _pct_or_none(klav["cumpl_prod"], 1)},
                {"Proceso": "Empaque",              "Plan": kemp["estandar_pct"] * 100, "Real": _pct_or_none(kemp["pct_empacado"], 1), "Unidad": "%",             "Cumplimiento %": _pct_or_none(kemp["cumpl_empacado"], 1)},
                {"Proceso": "Cargue y despacho",    "Plan": kcar["estandar_ton"],     "Real": _round_or_none(kcar["ton_promedio"], 2),  "Unidad": "t/despacho",         "Cumplimiento %": _pct_or_none(kcar["cumpl_ton"], 1)},
            ])
            prod.to_excel(writer, sheet_name="2. Productividad", index=False)

            # --- 3. Calidad ----------------------------------------------
            cal = pd.DataFrame([
                {"Indicador": "Primera calidad",      "Real %": _pct_or_none(klav["pct_1"], 1), "Objetivo %": 60,   "Total kg": klav["total_1"]},
                {"Indicador": "Segunda calidad",      "Real %": _pct_or_none(klav["pct_2"], 1), "Objetivo %": 25,   "Total kg": klav["total_2"]},
                {"Indicador": "Tercera calidad",      "Real %": _pct_or_none(klav["pct_3"], 1), "Objetivo %": 15,   "Total kg": klav["total_3"]},
                {"Indicador": "Caja comercializable", "Real %": _pct_or_none(klav["pct_comercial"], 1), "Objetivo %": 90,  "Total kg": klav["total_1"] + klav["total_2"] + klav["total_3"]},
                {"Indicador": "Semilla",              "Real %": _pct_or_none(klav["pct_semilla"], 1), "Objetivo %": "-", "Total kg": klav["total_semilla"]},
                {"Indicador": "Pérdida en lavado",    "Real %": _pct_or_none(klav["pct_perdida"], 2), "Objetivo %": "<5", "Total kg": klav["total_descarte"]},
                {"Indicador": "Descarte en cosecha",  "Real %": _pct_or_none(kcos["pct_descarte"], 2), "Objetivo %": "<5", "Total kg": kcos["total_descarte"]},
            ])
            cal.to_excel(writer, sheet_name="3. Calidad", index=False)

            # --- 4. Dotacion ---------------------------------------------
            dot_rows = []
            diag_map = {"subcontratacion": "Subcontratación",
                        "sobrecontratacion": "Sobrecontratación",
                        "alineado": "Alineado", "sin_datos": "Sin datos"}
            for p in PROCESOS:
                d = analisis_dotacion(data[p], p)
                dot_rows.append({
                    "Proceso":              PROCESOS_LABEL[p],
                    "Requerido (personas)": d["estandar"],
                    "Promedio asignado":    round(d["promedio_real"], 1) if d.get("promedio_real") is not None else None,
                    "Desviación %":         round(d["desviacion_pct"]*100, 1) if d.get("desviacion_pct") is not None else None,
                    "Días operativos":      d.get("total_jornadas", 0),
                    "Días subcontratados":  d.get("jornadas_sub", 0),
                    "Días sobrecontratados": d.get("jornadas_sobre", 0),
                    "Diagnóstico":          diag_map.get(d["diagnostico"], "-"),
                })
            pd.DataFrame(dot_rows).to_excel(writer, sheet_name="4. Dotacion", index=False)

            # --- 5. Plan vs Real -----------------------------------------
            pvr = pd.DataFrame([
                {"Proceso": "Corte", "Indicador": "Esquejes/persona/día",
                 "Plan": kcor["estandar_persona"],
                 "Real": _round_or_none(kcor["prod_persona_dia"]),
                 "Desv. absoluta": round(kcor["prod_persona_dia"] - kcor["estandar_persona"]) if kcor["prod_persona_dia"] is not None else None,
                 "Desv. %": round((kcor["cumpl_prod"] - 1)*100, 1) if kcor["cumpl_prod"] is not None else None},
                {"Proceso": "Siembra", "Indicador": "Plantas/persona/día",
                 "Plan": ksie["estandar_persona"],
                 "Real": _round_or_none(ksie["prod_persona_dia"]),
                 "Desv. absoluta": round(ksie["prod_persona_dia"] - ksie["estandar_persona"]) if ksie["prod_persona_dia"] is not None else None,
                 "Desv. %": round((ksie["cumpl_prod"] - 1)*100, 1) if ksie["cumpl_prod"] is not None else None},
                {"Proceso": "Cosecha", "Indicador": "Kg/persona/día",
                 "Plan": kcos["estandar_persona"],
                 "Real": _round_or_none(kcos["prod_persona_dia"]),
                 "Desv. absoluta": round(kcos["prod_persona_dia"] - kcos["estandar_persona"]) if kcos["prod_persona_dia"] is not None else None,
                 "Desv. %": round((kcos["cumpl_prod"] - 1)*100, 1) if kcos["cumpl_prod"] is not None else None},
                {"Proceso": "Cosecha", "Indicador": "Toneladas del período",
                 "Plan": meta_ton_periodo,
                 "Real": round(kcos["total_kg"]/1000, 2),
                 "Desv. absoluta": round((kcos["total_kg"]/1000) - meta_ton_periodo, 2) if meta_ton_periodo is not None else None,
                 "Desv. %": round((kcos["cumpl_meta_periodo"] - 1)*100, 1) if kcos["cumpl_meta_periodo"] is not None else None},
                {"Proceso": "Lavado", "Indicador": "Kg/persona/día",
                 "Plan": klav["estandar_persona"],
                 "Real": _round_or_none(klav["prod_persona_dia"]),
                 "Desv. absoluta": round(klav["prod_persona_dia"] - klav["estandar_persona"]) if klav["prod_persona_dia"] is not None else None,
                 "Desv. %": round((klav["cumpl_prod"] - 1)*100, 1) if klav["cumpl_prod"] is not None else None},
                {"Proceso": "Empaque", "Indicador": "% Empacado / recibido",
                 "Plan": kemp["estandar_pct"] * 100,
                 "Real": _pct_or_none(kemp["pct_empacado"], 1),
                 "Desv. absoluta": round((kemp["pct_empacado"] - kemp["estandar_pct"]) * 100, 1) if kemp["pct_empacado"] is not None else None,
                 "Desv. %": round((kemp["cumpl_empacado"] - 1) * 100, 1) if kemp["cumpl_empacado"] is not None else None},
                {"Proceso": "Cargue", "Indicador": "Toneladas/despacho",
                 "Plan": ESTANDARES["Cargue Vehiculo"]["toneladas"],
                 "Real": _round_or_none(kcar["ton_promedio"], 2),
                 "Desv. absoluta": round(kcar["ton_promedio"] - ESTANDARES["Cargue Vehiculo"]["toneladas"], 2) if kcar["ton_promedio"] is not None else None,
                 "Desv. %": round((kcar["cumpl_ton"] - 1)*100, 1) if kcar["cumpl_ton"] is not None else None},
            ])
            pvr.to_excel(writer, sheet_name="5. Plan vs Real", index=False)

            # --- 6. Detalle por proceso ----------------------------------
            for p in PROCESOS:
                dfp = data[p]
                if not dfp.empty:
                    excluir = {"Anio", "Mes", "Semana", "Origen Datos"}
                    cols = [c for c in dfp.columns if c not in excluir]
                    tmp = dfp[cols].copy()
                    for c in tmp.columns:
                        if pd.api.types.is_datetime64_any_dtype(tmp[c]):
                            try:
                                tmp[c] = tmp[c].dt.tz_localize(None) if tmp[c].dt.tz else tmp[c]
                            except Exception:
                                pass
                    safe = f"6. {PROCESOS_LABEL[p][:25]}"
                    tmp.to_excel(writer, sheet_name=safe[:31], index=False)

            # --- 7. Calidad de datos -------------------------------------
            data_all = store_to_dataframes(store_data)  # sin filtros
            inc = detectar_inconsistencias(data_all)
            if inc.empty:
                inc = pd.DataFrame([{"Resultado": "Sin inconsistencias detectadas"}])
            else:
                if "Fecha" in inc.columns:
                    inc["Fecha"] = pd.to_datetime(inc["Fecha"], errors="coerce")
                    try:
                        inc["Fecha"] = inc["Fecha"].dt.tz_localize(None) if inc["Fecha"].dt.tz is not None else inc["Fecha"]
                    except Exception:
                        pass
            inc.to_excel(writer, sheet_name="7. Calidad de datos", index=False)

            # --- 8. Parametros -------------------------------------------
            params_rows = []
            for proc, cfg in ESTANDARES.items():
                for kk, vv in cfg.items():
                    params_rows.append({"Categoria": "Estándar",
                                        "Proceso": PROCESOS_LABEL[proc],
                                        "Parámetro": kk.replace("_", " "),
                                        "Valor": vv})
            for proc, vv in DOTACION_ESTANDAR.items():
                params_rows.append({"Categoria": "Dotación",
                                    "Proceso": PROCESOS_LABEL[proc],
                                    "Parámetro": "personas requeridas",
                                    "Valor": vv})
            umbs = [
                ("Cumplimiento OK",        UMBRAL_CUMPLIMIENTO_OK),
                ("Cumplimiento alerta",    UMBRAL_CUMPLIMIENTO_ALERTA),
                ("Primera calidad OK",     UMBRAL_PRIMERA_OK),
                ("Primera calidad alerta", UMBRAL_PRIMERA_ALERTA),
                ("Comercializable OK",     UMBRAL_COMERCIAL_OK),
                ("Comercializable alerta", UMBRAL_COMERCIAL_ALERTA),
                ("Pérdida OK",             UMBRAL_PERDIDA_OK),
                ("Pérdida alerta",         UMBRAL_PERDIDA_ALERTA),
            ]
            for name, val in umbs:
                params_rows.append({"Categoria": "Umbral", "Proceso": "-",
                                    "Parámetro": name, "Valor": val})
            pd.DataFrame(params_rows).to_excel(writer, sheet_name="8. Parametros", index=False)

            # Formato corporativo
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                if ws.max_row < 1:
                    continue
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill("solid", fgColor="1F3A5F")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value)) for c in col if c.value is not None),
                                  default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
                ws.freeze_panes = "A2"

    return dcc.send_bytes(_write,
        f"puro_finca_consolidado_{date.today()}.xlsx")
